from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from ..config import settings


def _safe_stem(text: str) -> str:
    return "".join(character if character.isalnum() or character in {"-", "_"} else "_" for character in text)


def _safe_sheet_title(text: str) -> str:
    invalid = {'\\', '/', '*', '?', ':', '[', ']'}
    cleaned = "".join("_" if character in invalid else character for character in text).strip()
    return cleaned[:31] or "sheet"


def _voucher_tables(voucher: dict) -> list[dict]:
    document_json = voucher.get("document_json") or {}
    tables = document_json.get("tables") or []
    return [table for table in tables if isinstance(table, dict)]


def export_voucher_xlsx(voucher: dict) -> Path:
    workbook = Workbook()
    header_sheet = workbook.active
    header_sheet.title = "header"
    items_sheet = workbook.create_sheet("items")
    table_index_sheet = workbook.create_sheet("tables_index")

    header_sheet.append(
        [
            "voucher_id",
            "type",
            "status",
            "issue_date",
            "due_date",
            "document_number",
            "vendor_name",
            "customer_name",
            "currency",
            "subtotal",
            "tax",
            "discount",
            "grand_total",
            "notes",
            "needs_review",
        ]
    )
    header_sheet.append(
        [
            voucher["id"],
            voucher["type"],
            voucher["status"],
            voucher.get("issue_date"),
            voucher.get("due_date"),
            voucher.get("document_number"),
            voucher.get("vendor_name"),
            voucher.get("customer_name"),
            voucher.get("currency"),
            voucher.get("subtotal"),
            voucher.get("tax"),
            voucher.get("discount"),
            voucher.get("grand_total"),
            voucher.get("notes"),
            voucher.get("needs_review"),
        ]
    )

    items_sheet.append(
        [
            "voucher_id",
            "line_no",
            "description",
            "quantity",
            "unit",
            "unit_price",
            "amount",
            "tax_rate",
            "confidence",
            "needs_review",
        ]
    )
    for item in voucher["items"]:
        items_sheet.append(
            [
                voucher["id"],
                item.get("line_no"),
                item.get("description"),
                item.get("quantity"),
                item.get("unit"),
                item.get("unit_price"),
                item.get("amount"),
                item.get("tax_rate"),
                item.get("confidence"),
                item.get("needs_review"),
            ]
        )

    table_index_sheet.append(
        [
            "voucher_id",
            "page",
            "table_index",
            "sheet_name",
            "column_count",
            "row_count",
            "bbox_json",
        ]
    )
    for table in _voucher_tables(voucher):
        page = table.get("page")
        table_index = table.get("table_index")
        headers = [str(header or "") for header in (table.get("headers") or [])]
        rows = [
            [str(cell or "") for cell in row]
            for row in (table.get("rows") or [])
        ]
        sheet_name = _safe_sheet_title(f"table_p{page}_{table_index}")
        table_sheet = workbook.create_sheet(sheet_name)
        table_sheet.append(headers if headers else [""])
        for row in rows:
            width = max(len(headers), len(row))
            padded = row + [""] * (width - len(row))
            table_sheet.append(padded)

        table_index_sheet.append(
            [
                voucher["id"],
                page,
                table_index,
                sheet_name,
                len(headers),
                len(rows),
                json.dumps(table.get("bbox") or [], ensure_ascii=False),
            ]
        )

    filename = f"{_safe_stem(voucher['id'])}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    destination = settings.export_dir / filename
    workbook.save(destination)
    return destination


def export_voucher_csv_zip(voucher: dict) -> Path:
    header_buffer = io.StringIO()
    item_buffer = io.StringIO()
    table_index_buffer = io.StringIO()

    header_writer = csv.DictWriter(
        header_buffer,
        fieldnames=[
            "voucher_id",
            "type",
            "status",
            "issue_date",
            "due_date",
            "document_number",
            "vendor_name",
            "customer_name",
            "currency",
            "subtotal",
            "tax",
            "discount",
            "grand_total",
            "notes",
            "needs_review",
        ],
    )
    header_writer.writeheader()
    header_writer.writerow(
        {
            "voucher_id": voucher["id"],
            "type": voucher["type"],
            "status": voucher["status"],
            "issue_date": voucher.get("issue_date"),
            "due_date": voucher.get("due_date"),
            "document_number": voucher.get("document_number"),
            "vendor_name": voucher.get("vendor_name"),
            "customer_name": voucher.get("customer_name"),
            "currency": voucher.get("currency"),
            "subtotal": voucher.get("subtotal"),
            "tax": voucher.get("tax"),
            "discount": voucher.get("discount"),
            "grand_total": voucher.get("grand_total"),
            "notes": voucher.get("notes"),
            "needs_review": voucher.get("needs_review"),
        }
    )

    item_writer = csv.DictWriter(
        item_buffer,
        fieldnames=[
            "voucher_id",
            "line_no",
            "description",
            "quantity",
            "unit",
            "unit_price",
            "amount",
            "tax_rate",
            "confidence",
            "needs_review",
        ],
    )
    item_writer.writeheader()
    for item in voucher["items"]:
        item_writer.writerow(
            {
                "voucher_id": voucher["id"],
                "line_no": item.get("line_no"),
                "description": item.get("description"),
                "quantity": item.get("quantity"),
                "unit": item.get("unit"),
                "unit_price": item.get("unit_price"),
                "amount": item.get("amount"),
                "tax_rate": item.get("tax_rate"),
                "confidence": item.get("confidence"),
                "needs_review": item.get("needs_review"),
            }
        )

    table_index_writer = csv.DictWriter(
        table_index_buffer,
        fieldnames=[
            "voucher_id",
            "page",
            "table_index",
            "file_name",
            "column_count",
            "row_count",
            "bbox_json",
        ],
    )
    table_index_writer.writeheader()
    table_archives: list[tuple[str, bytes]] = []
    for table in _voucher_tables(voucher):
        page = table.get("page")
        table_index = table.get("table_index")
        headers = [str(header or "") for header in (table.get("headers") or [])]
        rows = [
            [str(cell or "") for cell in row]
            for row in (table.get("rows") or [])
        ]
        file_name = f"voucher_table_p{page}_{table_index}.csv"
        table_buffer = io.StringIO()
        writer = csv.writer(table_buffer)
        writer.writerow(headers if headers else [""])
        for row in rows:
            width = max(len(headers), len(row))
            padded = row + [""] * (width - len(row))
            writer.writerow(padded)
        table_archives.append((file_name, table_buffer.getvalue().encode("utf-8-sig")))
        table_index_writer.writerow(
            {
                "voucher_id": voucher["id"],
                "page": page,
                "table_index": table_index,
                "file_name": file_name,
                "column_count": len(headers),
                "row_count": len(rows),
                "bbox_json": json.dumps(table.get("bbox") or [], ensure_ascii=False),
            }
        )

    filename = f"{_safe_stem(voucher['id'])}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.zip"
    destination = settings.export_dir / filename
    with zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("voucher_header.csv", header_buffer.getvalue().encode("utf-8-sig"))
        archive.writestr("voucher_items.csv", item_buffer.getvalue().encode("utf-8-sig"))
        archive.writestr("voucher_tables_index.csv", table_index_buffer.getvalue().encode("utf-8-sig"))
        for file_name, payload in table_archives:
            archive.writestr(file_name, payload)
    return destination
